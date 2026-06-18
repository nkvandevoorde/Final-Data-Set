import googlemaps
import openpyxl
import time

#Initialize the Google Maps client and load the Excel file
gmaps = googlemaps.Client(key='AIzaSyBxqpE1S2uo7zAZy4zjKJRoYeJq6vHe8Lc')
data = openpyxl.load_workbook('Final_Data.xlsx')

#Get addresses from sheet
sheet = data.active
addresses = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    addresses.append(row[6]) #addresses in column G

#Geocode the addresses
geocoded_addresses = []
for address in addresses:
    if address:
        geocoded = gmaps.geocode(address)
        time.sleep(0.1) 
        if geocoded:
            geocoded_addresses.append(geocoded[0]['geometry']['location'])
        else:
            geocoded_addresses.append(None)
    else:
        geocoded_addresses.append(None)

# Save the geocoded addresses to the Excel file
sheet.cell(row=1, column=14).value = 'Latitude'
sheet.cell(row=1, column=15).value = 'Longitude'

for i, geocoded in enumerate(geocoded_addresses, start=2):
    if geocoded:
        sheet.cell(row=i, column=14).value = geocoded['lat']
        sheet.cell(row=i, column=15).value = geocoded['lng']
    else:
        sheet.cell(row=i, column=14).value = None
        sheet.cell(row=i, column=15).value = None

data.save('Final_Data.xlsx')
