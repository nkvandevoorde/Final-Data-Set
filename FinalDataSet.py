from openpyxl import Workbook
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl import load_workbook
import unicodedata
import pandas as pd
from rapidfuzz import fuzz
import re

cnes_wb = load_workbook("Final_CNES_data_Cleaned.xlsx").active
rel_wb = load_workbook("/Users/noravandevoorde/Downloads/SPHERE/Clean_Relatorio_Book.xlsx").active

# #Print column names in both datasets
# print("Columns in CNES dataset:", [cell.value for cell in cnes_wb[1] if cell.value is not None])
# print("Columns in Relatório dataset:", [cell.value for cell in rel_wb[1] if cell.value is not None])

##Relatório Data Processing

# #Concatenate the Relatório Number, Addresss, and Zip Code columns into one "Site Address" Column
# site_address_col = rel_wb.max_column+1
# rel_wb.cell(row=1, column=site_address_col).value = "Site Address"

# for excel_row, row in enumerate(rel_wb.iter_rows(min_row=2, values_only=True), start=2):
#     rel_number = row[8]  # Relatório number in column I
#     address = row[9]  # Address in column J
#     zip_code = row[10]  # Zip Code in column K
#     site_address = f"{rel_number} - {address} - {zip_code}"
#     rel_wb.cell(row=excel_row, column=site_address_col).value = site_address

# #Compare CEP and Zip Code columns
# highlight = PatternFill(fill_type="solid", fgColor="d17669")
# mismatch_count = 0

# for row in range(2, rel_wb.max_row + 1):
#     cep = str(rel_wb.cell(row=row, column=4).value).strip()  # CEP in column D
#     zip_code = str(rel_wb.cell(row=row, column=11).value).strip()  # Zip Code in column K
#     if cep != zip_code:
#         rel_wb.cell(row=row, column=4).fill = highlight
#         rel_wb.cell(row=row, column=11).fill = highlight
#         mismatch_count += 1

# print(f"Number of mismatches: {mismatch_count}")
# #519/2616 = 19.839% Zip Code Mismatches
# rel_wb.parent.save("Clean_Relatorio_Book.xlsx")

# #Comparing addresses in Relatório dataset
# def clean_address(address):
#     if address is None:
#         return ""
    
#     address = str(address).upper()

#     # Remove accents
#     address = ''.join(c for c in unicodedata.normalize('NFD', address) if unicodedata.category(c) != 'Mn')

#     #Remove Punctuation
#     address = re.sub(r'[^A-Z0-9\s]', ' ', address)

#     #Remove extra spaces
#     address = re.sub(r'\s+', ' ', address).strip()

#     return address

# orange = PatternFill(fill_type="solid", fgColor="c28651")
# mismatch_amt = 0

# for row in range(2, rel_wb.max_row + 1):
#     sheets_address = clean_address(rel_wb.cell(row=row, column=5).value)  # Address in column E
#     site_address = clean_address(rel_wb.cell(row=row, column=12).value)  # Site Address in column L

#     similarity = fuzz.token_set_ratio(sheets_address, site_address)

#     #Threshold
#     if similarity < 60:  # Adjust the threshold as needed
#         rel_wb.cell(row=row, column=5).fill = orange
#         rel_wb.cell(row=row, column=12).fill = orange
#         mismatch_amt += 1

# print(f"Number of address mismatches: {mismatch_amt}")
# #830/2616 = 31.72% (roughly)Address Mismatches
# rel_wb.parent.save("Clean_Relatorio_Book.xlsx")

#Check for matching CNES numbers between the two datasets
#Extract CNES and Relatório numbers
cnes_numbers = [str(cell.value).strip() for cell in cnes_wb['C'][1:] if cell.value is not None] #CNES in column C
rel_numbers = [str(cell.value).strip() for cell in rel_wb['B'][1:] if cell.value is not None] #Relatório in column B
#Find matches and outliers
matching_numbers = set(cnes_numbers) & set(rel_numbers)
cnes_outliers = set(cnes_numbers) - set(rel_numbers)
rel_outliers = set(rel_numbers) - set(cnes_numbers)
#print("Relatório outliers:", rel_outliers)
print("Number of matching CNES numbers:", len(matching_numbers))
print("Number of CNES outliers:", len(cnes_outliers))
print("Number of Relatório outliers:", len(rel_outliers))

#Create Library for CNES sheet
cnes_library = {}
for excel_row, row in enumerate(cnes_wb.iter_rows(min_row=2, values_only=True), start=2):
    cnes_library[excel_row] = {
        "State": str(row[0]).strip() if row[0] is not None else "",  # State in column A
        "City": str(row[1]).strip() if row[1] is not None else "",   # City in column B
        "CNES": str(row[2]).strip() if row[2] is not None else "",   # CNES in column C
        "Latitude (CNES)": str(row[3]).strip() if row[3] is not None else "", # Latitude in column D
        "Longitude (CNES)": str(row[4]).strip() if row[4] is not None else "", # Longitude in column E
        "REGIC Label": str(row[5]).strip() if row[5] is not None else "", # REGIC Label in column F
        "Address": str(row[6]).strip() if row[6] is not None else "", # Address in column G
        }
#print(list(cnes_library.items())[:5])

#Create Library for Relatório sheet
rel_library = {}
for excel_row, row in enumerate(rel_wb.iter_rows(min_row=2, values_only=True), start=2):
        rel_library[excel_row] = {
            "State": str(row[0]).strip() if row[0] is not None else "",  # State in column A
            "CNES": str(row[1]).strip() if row[1] is not None else "",   # CNES in column B
            "City": str(row[2]).strip() if row[2] is not None else "",   # City in column C
            "ZIP Code (sheet)": str(row[3]).strip() if row[3] is not None else "", # ZIP Code in column D
            "Address (sheet)": str(row[4]).strip() if row[4] is not None else "", # Address in column E
            "Latitude (Relatorio)": str(row[6]).strip() if row[6] is not None else "", # Latitude in column G
            "Longitude (Relatorio)": str(row[7]).strip() if row[7] is not None else "", # Longitude in column H
            "ZIP Code (site)": str(row[10]).strip() if row[10] is not None else "", # ZIP Code in column K
            "Address (site)": str(row[11]).strip() if row[11] is not None else "", # Address in column L
        }

#print(list(rel_library.items())[:5])

#Create CNES matches library
cnes_matches = {}
for cnes_row, cnes_data in cnes_library.items():
    for rel_row, rel_data in rel_library.items():
        if str(cnes_data["CNES"]) == str(rel_data["CNES"]):
            cnes_matches[cnes_row] = {
                "CNES": str(cnes_data["CNES"]) if cnes_data["CNES"] is not None else "",
                "State": str(cnes_data["State"]) if cnes_data["State"] is not None else "",
                "City": str(cnes_data["City"]) if cnes_data["City"] is not None else "",
                "ZIP Code (sheet)": str(rel_data["ZIP Code (sheet)"]) if rel_data["ZIP Code (sheet)"] is not None else "",
                "ZIP Code (site)": str(rel_data["ZIP Code (site)"]) if rel_data["ZIP Code (site)"] is not None else "",
                "Address (sheet)": str(rel_data["Address (sheet)"]) if rel_data["Address (sheet)"] is not None else "",
                "Address (site)": str(rel_data["Address (site)"]) if rel_data["Address (site)"] is not None else "",
                "Latitude (Relatorio)": str(rel_data["Latitude (Relatorio)"]) if rel_data["Latitude (Relatorio)"] is not None else "",
                "Longitude (Relatorio)": str(rel_data["Longitude (Relatorio)"]) if rel_data["Longitude (Relatorio)"] is not None else "",
                "Latitude (CNES)": str(cnes_data["Latitude (CNES)"]) if cnes_data["Latitude (CNES)"] is not None else "",
                "Longitude (CNES)": str(cnes_data["Longitude (CNES)"]) if cnes_data["Longitude (CNES)"] is not None else "",
                "REGIC Label": str(cnes_data["REGIC Label"]) if cnes_data["REGIC Label"] is not None else ""
            }
            break
#print(list(cnes_matches.items())[:5])
print("Number of CNES matches:", len(cnes_matches))

#Create library for CNES outliers
cnes_outliers_lib = {}
for cnes_row, cnes_data in cnes_library.items():
    if str(cnes_data["CNES"]) not in [str(match["CNES"]) for match in cnes_matches.values()]:
        cnes_outliers_lib[cnes_row] = {
            "CNES": str(cnes_data["CNES"]) if cnes_data["CNES"] is not None else "",
            "State": str(cnes_data["State"]) if cnes_data["State"] is not None else "",
            "City": str(cnes_data["City"]) if cnes_data["City"] is not None else "",
            "Address (site)": str(cnes_data["Address"]) if cnes_data["Address"] is not None else "",
            "Latitude (CNES)": str(cnes_data["Latitude (CNES)"]) if cnes_data["Latitude (CNES)"] is not None else "",
            "Longitude (CNES)": str(cnes_data["Longitude (CNES)"]) if cnes_data["Longitude (CNES)"] is not None else "",
            "REGIC Label": str(cnes_data["REGIC Label"]) if cnes_data["REGIC Label"] is not None else ""
        }

#print(list(cnes_outliers_lib.items())[:5])
print("Number of CNES outliers:", len(cnes_outliers_lib))

#Create library for Relatório outliers
rel_outliers_lib = {}
for rel_row, rel_data in rel_library.items():
    if str(rel_data["CNES"]) not in [str(match["CNES"]) for match in cnes_matches.values()]:
        rel_outliers_lib[rel_row] = {
            "CNES": str(rel_data["CNES"]) if rel_data["CNES"] is not None else "",
            "State": str(rel_data["State"]) if rel_data["State"] is not None else "",
            "City": str(rel_data["City"]) if rel_data["City"] is not None else "",
            "ZIP Code (sheet)": str(rel_data["ZIP Code (sheet)"]) if rel_data["ZIP Code (sheet)"] is not None else "",
            "ZIP Code (site)": str(rel_data["ZIP Code (site)"]) if rel_data["ZIP Code (site)"] is not None else "",
            "Address (sheet)": str(rel_data["Address (sheet)"]) if rel_data["Address (sheet)"] is not None else "",
            "Address (site)": str(rel_data["Address (site)"]) if rel_data["Address (site)"] is not None else "",
            "Latitude (Relatorio)": str(rel_data["Latitude (Relatorio)"]) if rel_data["Latitude (Relatorio)"] is not None else "",
            "Longitude (Relatorio)": str(rel_data["Longitude (Relatorio)"]) if rel_data["Longitude (Relatorio)"] is not None else ""
        }

#print(list(rel_outliers_lib.items())[:5])
print("Number of Relatório outliers:", len(rel_outliers_lib))

#Load libraries into the columns of the new excel file
#Create new sheet to combine relevant data
new_wb = Workbook()
new_ws = new_wb.active
new_ws.title = "Final Data"
new_ws.append(["CNES", "State", "City", "ZIP Code (sheet)", "ZIP Code (site)", "Address(sheet)", "Address(site)", "Latitude (Relatorio)", 
                  "Longitude (Relatorio)", "Latitude (CNES)", "Longitude (CNES)", "REGIC Label"])
new_wb.save("Final_Data.xlsx")
current_rows = new_ws.max_row + 1

for library in [cnes_matches, cnes_outliers_lib, rel_outliers_lib]:
    for _, data in library.items():
        new_ws.cell(row=current_rows, column=1).value = data.get("CNES", "")
        new_ws.cell(row=current_rows, column=2).value = data.get("State", "")
        new_ws.cell(row=current_rows, column=3).value = data.get("City", "")
        new_ws.cell(row=current_rows, column=4).value = data.get("ZIP Code (sheet)", "")
        new_ws.cell(row=current_rows, column=5).value = data.get("ZIP Code (site)", "")
        new_ws.cell(row=current_rows, column=6).value = data.get("Address (sheet)", "")
        new_ws.cell(row=current_rows, column=7).value = data.get("Address (site)", "")
        new_ws.cell(row=current_rows, column=8).value = data.get("Latitude (Relatorio)", "")
        new_ws.cell(row=current_rows, column=9).value = data.get("Longitude (Relatorio)", "")
        new_ws.cell(row=current_rows, column=10).value = data.get("Latitude (CNES)", "")
        new_ws.cell(row=current_rows, column=11).value = data.get("Longitude (CNES)", "")
        new_ws.cell(row=current_rows, column=12).value = data.get("REGIC Label", "")
        current_rows += 1

new_wb.save("Final_Data.xlsx")

#Compare CEP and Zip Code columns
highlight = PatternFill(fill_type="solid", fgColor="d17669")
mismatch_count = 0

for row in range(2, new_ws.max_row + 1):
    cep = str(new_ws.cell(row=row, column=4).value).strip()  # CEP in column D
    zip_code = str(new_ws.cell(row=row, column=5).value).strip()  # Zip Code in column E
    if cep != zip_code:
        new_ws.cell(row=row, column=4).fill = highlight
        new_ws.cell(row=row, column=5).fill = highlight
        mismatch_count += 1

print(f"Number of mismatches: {mismatch_count}")
#511/2616 = 19.839% Zip Code Mismatches
new_wb.save("Final_Data.xlsx")

#Comparing addresses in Relatório dataset
def clean_address(address):
    if address is None:
        return ""
    
    address = str(address).upper()

    # Remove accents
    address = ''.join(c for c in unicodedata.normalize('NFD', address) if unicodedata.category(c) != 'Mn')

    #Remove Punctuation
    address = re.sub(r'[^A-Z0-9\s]', ' ', address)

    #Remove extra spaces
    address = re.sub(r'\s+', ' ', address).strip()

    return address

orange = PatternFill(fill_type="solid", fgColor="c28651")
mismatch_amt = 0

for row in range(2, new_ws.max_row + 1):
    sheets_address = clean_address(new_ws.cell(row=row, column=6).value)  # Address in column F
    site_address = clean_address(new_ws.cell(row=row, column=7).value)  # Site Address in column G

    if not sheets_address or not site_address:
        continue

    similarity = fuzz.token_set_ratio(sheets_address, site_address)

    #Threshold
    if similarity < 60:  # Adjust the threshold as needed
        new_ws.cell(row=row, column=6).fill = orange
        new_ws.cell(row=row, column=7).fill = orange
        mismatch_amt += 1

print(f"Number of address mismatches: {mismatch_amt}")
#830/2616 = 31.72% (roughly)Address Mismatches
new_wb.save("Final_Data.xlsx")

#Compare latitudes and longitudes using tolerance
def compare_coordinates(lat1, lng1, lat2, lng2, tolerance=0.001):
   try:
       lat1 = float(str(lat1).strip())
       lng1 = float(str(lng1).strip())
       lat2 = float(str(lat2).strip())
       lng2 = float(str(lng2).strip())
   except (ValueError, TypeError):
       return False

   return (abs(lat1 - lat2) <= tolerance) and (abs(lng1 - lng2) <= tolerance)

blue = PatternFill(fill_type="solid", fgColor="FFA4ECD6")
mismatch_count = 0

for row in range(2, new_ws.max_row + 1):
    rel_lat = new_ws.cell(row=row, column=8).value  # Latitude in column H
    rel_lng = new_ws.cell(row=row, column=9).value  # Longitude in column I
    cnes_lat = new_ws.cell(row=row, column=10).value  # Latitude in column J
    cnes_lng = new_ws.cell(row=row, column=11).value  # Longitude in column K

    if not rel_lat or not rel_lng or not cnes_lat or not cnes_lng:
        continue

    if not compare_coordinates(rel_lat, rel_lng, cnes_lat, cnes_lng):
        new_ws.cell(row=row, column=8).fill = blue
        new_ws.cell(row=row, column=9).fill = blue
        new_ws.cell(row=row, column=10).fill = blue
        new_ws.cell(row=row, column=11).fill = blue
        mismatch_count += 1

print(f"Number of coordinate mismatches: {mismatch_count}")
#3851/4168 = 92.38% Coordinate Mismatches -> can change the tolerancce depending on how accurate the coordinates need to be
new_wb.save("Final_Data.xlsx")

#Grab remaining REGIC labels for Relatorio dataset
#Remove special characters for consistent spelling
def remove_accents(input_str):
    if input_str is None:
        return None
    nfkd_form = unicodedata.normalize('NFKD', input_str)
    return ''.join([c for c in nfkd_form if not unicodedata.combining(c)])

#normalize the text
def clean(value):
    if value is None:
        return None
    return remove_accents(str(value).lower().strip())

#load REGIC data
regic_wb = load_workbook("/Users/noravandevoorde/Downloads/SPHERE/REGIC_COPY_0509.xlsx").active
regic_lib = {}
for excel_row, row in enumerate(regic_wb.iter_rows(min_row=2, values_only=True), start=2):
    regic_lib[excel_row] = {
        "State": clean(row[1]) if row[1] is not None else "",  # State in column B
        "City": clean(row[5]) if row[5] is not None else "",   # City in column F
        "REGIC_Label": clean(row[13]) if row[13] is not None else "", # REGIC Label in column N
        }
print(list(regic_lib.items())[:5])

#create a label lookup dictionary
regic_lookup = {}
for _, regic_data in regic_lib.items():
    key = (clean(regic_data["State"]), clean(regic_data["City"]))
    regic_lookup[key] = regic_data["REGIC_Label"]

#Compare state and city to rel_outlier_lib and add missing REGIC labels
city_state_matches = 0

for row in range(2, new_ws.max_row + 1):
    state = clean(new_ws.cell(row=row, column=2).value)  # State in column B
    city = clean(new_ws.cell(row=row, column=3).value)   # City in column C

    key = (state, city)
    current_label = new_ws.cell(row=row, column=12).value

    if (not current_label or str(current_label).strip() == "") and (key in regic_lookup):
        new_ws.cell(row=row, column=12).value = regic_lookup[key]  # REGIC Label in column L
        city_state_matches += 1

print(f"Number of city-state matches: {city_state_matches}")
new_wb.save("Final_Data.xlsx")

#Make a state and regions key to add regions to the data
regions ={
    "RO": "North",
    "AC": "North",
    "AM": "North",
    "RR": "North",
    "PA": "North",
    "AP": "North",
    "TO": "North",
    "MA": "Northeast",
    "PI": "Northeast",
    "CE": "Northeast",
    "RN": "Northeast",
    "PB": "Northeast",
    "PE": "Northeast",
    "AL": "Northeast",
    "SE": "Northeast",
    "BA": "Northeast",
    "MG": "Southeast",
    "SP": "Southeast",
    "RJ": "Southeast",
    "ES": "Southeast",
    "RS": "South",
    "SC": "South",
    "PR": "South",
    "MT" : "Center-West",
    "GO" : "Center-West",
    "DF" : "Center-West",
    "MS" : "Center-West",
}

#Assign regions to the data
next_col = new_ws.max_column + 1
new_ws.cell(row=1, column=next_col).value = "Region"

state_count = 0
for row in range(2, new_ws.max_row + 1):
    state = new_ws.cell(row=row, column=2).value  #State in column B
    if state in regions:
        new_ws.cell(row=row, column=next_col).value = regions[state]  #Region in the new column
        state_count += 1

print(f"Number of states with assigned regions: {state_count}")
new_wb.save("Final_Data.xlsx")