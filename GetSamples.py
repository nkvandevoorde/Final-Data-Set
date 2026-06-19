import openpyxl
from openpyxl import Workbook
import random

#load the Excel file
full_data = openpyxl.load_workbook('/Users/noravandevoorde/Downloads/SPHERE/Final_Dataset_Geocoded_copy.xlsx')
sheet = full_data.active

#create a dictionary organized by label
label_dict = {}
for row in sheet.iter_rows(min_row=2, values_only=True):
    label = row[11]  #Labels in column L
    if label not in label_dict:
        label_dict[label] = []
    label_dict[label].append(row)

#map labels to numbers
label_numbers = {}

for label in label_dict.keys():
    if label and label[0].isdigit():  # checks if the first character is a number
        label_numbers[label] = int(label[0])

#create a sample dictionary
sample_dict = {}

for label, rows in label_dict.items():
    if label in label_numbers:
        number = label_numbers[label]
        
        if number not in sample_dict:
            sample_dict[number] = []

        sample_dict[number].extend(rows)

#grab a sample from each label
sample_frac = 0.1  # 10% sample

for label, rows in sample_dict.items():
    sample_size = int(len(rows) * sample_frac)
    sample_dict[label] = random.sample(rows, sample_size)

#store samples in sheets for validation
output = Workbook()
output.remove(output.active) 

for label, rows in sample_dict.items():
    ws = output.create_sheet(f"Label_{label}")
    for row in rows:
        ws.append(row)

output.save("REGIC_random_samples.xlsx")